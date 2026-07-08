import csv
import itertools
import json
import re
from decimal import Decimal

from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db import transaction
from django.db.models import ProtectedError
from django.db.models import Min, Sum
from django.db.models.functions import Coalesce
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse_lazy
from django.utils.text import slugify
from django.views import View
from django.views.generic import CreateView, DetailView, TemplateView

from accounts.models import Account
from billing.services import estimate_output_access
from projects.models import Estimate

from .forms import AssemblyForm, CalculationRuleFormSet, FormulaForm, ManualLineItemForm, MaterialForm
from .models import Assembly, CalculationRule, Formula, LineItem

CATEGORY_DEFAULT_ORDER = list(Assembly.Category.values)
CATEGORY_LABELS = dict(Assembly.Category.choices)
VALID_CATEGORIES = set(Assembly.Category.values)


def _effective_category():
    """A global assembly's classification can be edited after LineItems
    already exist with the old category baked in (LineItem.category is
    denormalized, since calculation_rule is SET_NULL and manual lines have no
    assembly at all) - Coalesce prefers the live assembly category so an
    after-the-fact fix shows up without regenerating every LineItem, falling
    back to the stored column for manual/orphaned rows. Both the Detail view
    and the CSV rollup use this so they never disagree."""
    return Coalesce('calculation_rule__assembly__category', 'category')


def _category_order_for(account):
    """The account's preferred group order, falling back to (and filling any
    gaps from) the doc's default order - a category with no saved preference
    yet (new account, or a category introduced since the account last
    reordered) never silently disappears from the page."""
    saved = [c for c in (account.category_order or []) if c in VALID_CATEGORIES]
    saved += [c for c in CATEGORY_DEFAULT_ORDER if c not in saved]
    return saved


def _item_rank(account, category, role):
    """Position of `role` within its category's saved item order, matched
    case/whitespace-insensitively - the seed data itself has near-duplicate
    role strings across assemblies (e.g. "Bottom Plate" vs "Bottom Plate
    (PT)") that won't share a saved rank without this. Unranked roles sort
    last, in whatever order they were already in."""
    order = (account.item_order or {}).get(category, [])
    normalized = [r.strip().lower() for r in order]
    key = (role or '').strip().lower()
    return normalized.index(key) if key in normalized else len(normalized)


def _grouped_order_list(estimate, page_id=None, page_only=False):
    """The supplier-ready view: line items grouped by product + piece length
    *and* construction system (so the same SKU used under two different
    systems, e.g. 2x6 SPF #2 in both walls and blocking, never merges into
    one row) with summed quantities, ordered to match the Detail page."""
    account = estimate.project.account
    category_rank = {c: i for i, c in enumerate(_category_order_for(account))}
    rows = list(
        _summary_line_items(estimate, page_id=page_id, page_only=page_only)
        .order_by()
        .values('effective_category', 'material_id', 'material__name', 'material__nominal_dimension',
                'material__species', 'material__grade', 'length_ft')
        .annotate(total_quantity=Sum('quantity'), min_role=Min('role'))
    )
    # min_role is a deterministic tiebreak for the rare case where a merged
    # row spans more than one role within the same category - not a source
    # of truth for item order at that granularity.
    rows.sort(key=lambda r: (
        category_rank.get(r['effective_category'], len(category_rank)),
        _item_rank(account, r['effective_category'], r['min_role']),
        r['material__name'],
    ))
    for row in rows:
        row['category_label'] = CATEGORY_LABELS.get(row['effective_category'], row['effective_category'])
        row['role_label'] = row.get('min_role') or ''
    return rows


def _attach_pricing(order_list, account):
    """Attaches the account's unit cost and the extended cost (unit x quantity)
    to each order row, and returns whether ANY row is priced. Rows without a
    price get unit_cost=None and extended_cost=None, so the template shows a
    dash and the estimate still reads as a clean material list. Never touches
    the global catalog: prices are read from the account's MaterialPrice rows."""
    from catalog.models import MaterialPrice

    material_ids = {row['material_id'] for row in order_list if row.get('material_id')}
    prices = dict(
        MaterialPrice.objects
        .filter(account=account, material_id__in=material_ids)
        .values_list('material_id', 'unit_cost')
    )
    any_priced = False
    for row in order_list:
        unit_cost = prices.get(row.get('material_id'))
        row['unit_cost'] = unit_cost
        if unit_cost is not None:
            row['extended_cost'] = unit_cost * (row['total_quantity'] or 0)
            any_priced = True
        else:
            row['extended_cost'] = None
    return any_priced


def _attach_label_lengths(order_list):
    """Rows whose line items carry no piece length (e.g. studs traced without
    a wall height in the settings snapshot) still deserve a usable label for
    the viewer's trace linking. Fall back to the measured run length of the
    linked traces when every linked trace agrees on one length. Display-only:
    the order list's length column stays honest (blank means uncut/unknown)."""
    from plans.geometry import measure_geometry
    from plans.models import Trace

    pending = [row for row in order_list if not row.get('length_ft') and row.get('trace_ids')]
    trace_ids = {trace_id for row in pending for trace_id in row['trace_ids']}
    if not trace_ids:
        return
    traces = {
        trace.pk: trace
        for trace in Trace.objects.filter(pk__in=trace_ids).select_related('plan_page')
    }
    for row in pending:
        lengths = set()
        for trace_id in row['trace_ids']:
            trace = traces.get(trace_id)
            if trace is None or trace.plan_page.scale_pixels_per_foot is None:
                continue
            measurement = measure_geometry(
                trace.tool_type, trace.geometry, trace.plan_page.scale_pixels_per_foot, trace.settings,
            )
            if 'length_ft' in measurement:
                lengths.add(int(round(float(measurement['length_ft']))))
        if len(lengths) == 1:
            row['label_length_ft'] = lengths.pop()


def _summary_line_items(estimate, page_id=None, page_only=False):
    line_items = estimate.line_items.annotate(effective_category=_effective_category())
    if page_only and page_id:
        line_items = line_items.filter(trace__plan_page_id=page_id)
    return line_items


def _attach_trace_context(order_list, estimate, current_page_id=None, page_only=False):
    """Add current-page and all-page trace context for each grouped summary
    row so the viewer can link material lines back to visible plan elements
    without losing awareness that a grouped row may span other pages."""
    total_trace_map = {}
    visible_trace_map = {}
    page_map = {}
    line_items = (
        _summary_line_items(estimate, page_id=current_page_id, page_only=page_only)
        .exclude(trace_id__isnull=True)
        .values(
            'effective_category', 'material__name', 'material__nominal_dimension',
            'material__species', 'material__grade', 'length_ft', 'trace_id', 'trace__plan_page_id',
        )
    )
    for item in line_items:
        key = (
            item['effective_category'],
            item['material__name'],
            item['material__nominal_dimension'],
            item['material__species'],
            item['material__grade'],
            item['length_ft'],
        )
        total_trace_map.setdefault(key, set()).add(item['trace_id'])
        page_map.setdefault(key, set()).add(item['trace__plan_page_id'])
        if current_page_id and item['trace__plan_page_id'] == current_page_id:
            visible_trace_map.setdefault(key, set()).add(item['trace_id'])
    for row in order_list:
        key = (
            row['effective_category'],
            row['material__name'],
            row['material__nominal_dimension'],
            row['material__species'],
            row['material__grade'],
            row['length_ft'],
        )
        total_trace_ids = sorted(total_trace_map.get(key, set()))
        visible_trace_ids = sorted(visible_trace_map.get(key, set())) if current_page_id else total_trace_ids
        row['trace_ids'] = visible_trace_ids
        row['visible_trace_count'] = len(visible_trace_ids)
        row['total_trace_count'] = len(total_trace_ids)
        row['page_count'] = len(page_map.get(key, set()))
    return order_list


# Plain "NxM" (2x6, 2 x 10, 1.75x11.875): dimensional lumber and engineered
# beam stock. Three-part sheet dims (7/16x4x8), rolls, and per-square siding
# do not match, so they are excluded from board feet rather than guessed at.
_DIMENSIONAL_RE = re.compile(r'^\s*(\d+(?:\.\d+)?)\s*[xX]\s*(\d+(?:\.\d+)?)\s*$')


def _summary_totals(order_list):
    """Headline numbers for the viewer's live summary bar: total pieces across
    every row, the number of distinct order rows, and framing board feet.
    BF = (thickness_in x width_in x length_ft) / 12 per piece, using the
    nominal dimension (a 2x6 counts as 2 x 6), only for rows whose dimension
    parses as plain NxM and that carry a piece length."""
    total_pieces = 0
    framing_bf = Decimal('0')
    material_cost = Decimal('0')
    priced_rows = 0
    for row in order_list:
        quantity = row['total_quantity'] or 0
        total_pieces += quantity
        match = _DIMENSIONAL_RE.match(row.get('material__nominal_dimension') or '')
        length_ft = row.get('length_ft')
        if match and length_ft:
            thickness = Decimal(match.group(1))
            width = Decimal(match.group(2))
            framing_bf += thickness * width * Decimal(length_ft) / 12 * quantity
        extended = row.get('extended_cost')
        if extended is not None:
            material_cost += extended
            priced_rows += 1
    return {
        'total_pieces': total_pieces,
        'row_count': len(order_list),
        'framing_bf': framing_bf,
        'material_cost': material_cost,
        'priced_rows': priced_rows,
        'has_pricing': priced_rows > 0,
    }


class EstimateDetailView(LoginRequiredMixin, DetailView):
    model = Estimate
    template_name = 'estimating/estimate_detail.html'
    context_object_name = 'estimate'

    def get_queryset(self):
        return Estimate.objects.for_account(self.request.user.account)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        account = self.request.user.account
        line_items = list(
            self.object.line_items.select_related('material', 'trace__plan_page')
            .annotate(effective_category=_effective_category())
            .order_by()
        )
        category_rank = {c: i for i, c in enumerate(_category_order_for(account))}
        line_items.sort(key=lambda li: (
            category_rank.get(li.effective_category, len(category_rank)),
            _item_rank(account, li.effective_category, li.role),
        ))
        grouped_line_items = [
            {'key': category, 'label': CATEGORY_LABELS.get(category, category), 'items': list(items)}
            for category, items in itertools.groupby(line_items, key=lambda li: li.effective_category)
        ]
        context['grouped_line_items'] = grouped_line_items
        order_list = _grouped_order_list(self.object)
        _attach_pricing(order_list, account)
        context['order_list'] = order_list
        context['totals'] = _summary_totals(order_list)
        context['manual_form'] = ManualLineItemForm(account=account)
        context['estimate_access'] = estimate_output_access(self.object)
        return context


class EstimateMaterialSummaryView(LoginRequiredMixin, DetailView):
    """Renders a compact, view-only material list partial - reuses the exact
    same _grouped_order_list() rollup as the full Estimate Detail page and
    CSV export, so the numbers can never disagree. Polled by the plan viewer
    after every trace create/update/delete to show a live-updating list next
    to the canvas without a full page reload; no drag-and-drop reordering
    here, that's what "Full view" (a link to the real Estimate Detail page)
    is for."""

    model = Estimate
    template_name = 'estimating/_material_summary.html'
    context_object_name = 'estimate'

    def get_queryset(self):
        return Estimate.objects.for_account(self.request.user.account)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        current_page_id = self.request.GET.get('current_page_id')
        page_only = self.request.GET.get('page_only') == '1'
        try:
            current_page_id = int(current_page_id) if current_page_id else None
        except (TypeError, ValueError):
            current_page_id = None
        order_list = _grouped_order_list(self.object, page_id=current_page_id, page_only=page_only)
        context['order_list'] = _attach_trace_context(
            order_list,
            self.object,
            current_page_id=current_page_id,
            page_only=page_only,
        )
        _attach_label_lengths(context['order_list'])
        _attach_pricing(context['order_list'], self.object.project.account)
        context['totals'] = _summary_totals(context['order_list'])
        context['page_only'] = page_only
        context['current_page_id'] = current_page_id
        return context


class EstimateCsvExportView(LoginRequiredMixin, View):
    """Download the grouped order list as CSV - the core deliverable."""

    def get(self, request, pk):
        estimate = get_object_or_404(Estimate.objects.for_account(request.user.account), pk=pk)
        access = estimate_output_access(estimate)
        if not access['can_access']:
            messages.info(
                request,
                'Export is unlocked per estimate or through a subscription. Choose an option below to continue.',
            )
            return redirect('estimating:estimate-detail', pk=estimate.pk)
        response = HttpResponse(content_type='text/csv')
        filename = f'{estimate.project.name} - {estimate.name} - materials.csv'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        writer = csv.writer(response)
        order_list = _grouped_order_list(estimate)
        has_pricing = _attach_pricing(order_list, estimate.project.account)
        header = ['System', 'Material', 'Dimension', 'Species/Grade', 'Length (ft)', 'Quantity']
        if has_pricing:
            header += ['Unit Cost', 'Extended Cost']
        writer.writerow(header)
        for row in order_list:
            species_grade = ' '.join(part for part in (row['material__species'], row['material__grade']) if part)
            line = [
                row['category_label'],
                row['material__name'],
                row['material__nominal_dimension'],
                species_grade,
                row['length_ft'] or '',
                row['total_quantity'],
            ]
            if has_pricing:
                line += [
                    '' if row['unit_cost'] is None else f'{row["unit_cost"]:.2f}',
                    '' if row['extended_cost'] is None else f'{row["extended_cost"]:.2f}',
                ]
            writer.writerow(line)
        if has_pricing:
            totals = _summary_totals(order_list)
            writer.writerow([])
            writer.writerow(['', '', '', '', '', 'Material total', '', f'{totals["material_cost"]:.2f}'])
        return response


class EstimatePrintView(LoginRequiredMixin, DetailView):
    model = Estimate
    template_name = 'estimating/estimate_print.html'
    context_object_name = 'estimate'

    def get_queryset(self):
        return Estimate.objects.for_account(self.request.user.account)

    def dispatch(self, request, *args, **kwargs):
        self.object = self.get_object()
        access = estimate_output_access(self.object)
        if not access['can_access']:
            messages.info(
                request,
                'Print-friendly output is unlocked per estimate or through a subscription.',
            )
            return redirect('estimating:estimate-detail', pk=self.object.pk)
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        order_list = _grouped_order_list(self.object)
        _attach_pricing(order_list, self.object.project.account)
        context['order_list'] = order_list
        context['totals'] = _summary_totals(order_list)
        context['estimate_access'] = estimate_output_access(self.object)
        return context


class CategoryOrderUpdateView(LoginRequiredMixin, View):
    """Saves the requesting account's preferred material-list group order.
    Account-wide by design - applies to every estimate the account has, past
    and future, the next time it's viewed."""

    def post(self, request):
        try:
            payload = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON.'}, status=400)
        order = payload.get('order')
        if not isinstance(order, list) or not all(isinstance(c, str) for c in order):
            return JsonResponse({'error': 'order must be a list of category keys.'}, status=400)
        unknown = [c for c in order if c not in VALID_CATEGORIES]
        if unknown:
            return JsonResponse({'error': f'Unknown category key(s): {", ".join(unknown)}.'}, status=400)
        with transaction.atomic():
            account = Account.objects.select_for_update().get(pk=request.user.account_id)
            account.category_order = order
            account.save(update_fields=['category_order'])
        return JsonResponse({'category_order': order})


class ItemOrderUpdateView(LoginRequiredMixin, View):
    """Saves the requesting account's preferred item order within one
    material-list group, keyed by role text (not by LineItem/CalculationRule
    id) so the preference generalizes to future estimates, per the
    account-wide requirement."""

    def post(self, request):
        try:
            payload = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON.'}, status=400)
        category = payload.get('category')
        order = payload.get('order')
        if category not in VALID_CATEGORIES:
            return JsonResponse({'error': 'Invalid category.'}, status=400)
        if not isinstance(order, list) or not all(isinstance(r, str) for r in order):
            return JsonResponse({'error': 'order must be a list of role strings.'}, status=400)
        with transaction.atomic():
            account = Account.objects.select_for_update().get(pk=request.user.account_id)
            item_order = dict(account.item_order or {})
            item_order[category] = order
            account.item_order = item_order
            account.save(update_fields=['item_order'])
        return JsonResponse({'category': category, 'order': order})


class ResetLayoutPreferencesView(LoginRequiredMixin, View):
    """Clears both saved layout preferences so a user can't get stuck after a
    confusing drag - back to the doc's default order."""

    def post(self, request):
        with transaction.atomic():
            account = Account.objects.select_for_update().get(pk=request.user.account_id)
            account.category_order = []
            account.item_order = {}
            account.save(update_fields=['category_order', 'item_order'])
        messages.success(request, 'Material list order reset to default.')
        return redirect('estimating:estimate-detail', pk=request.POST.get('estimate_id'))


class ManualLineItemCreateView(LoginRequiredMixin, View):
    def post(self, request, estimate_id):
        estimate = get_object_or_404(Estimate.objects.for_account(request.user.account), pk=estimate_id)
        form = ManualLineItemForm(request.POST, account=request.user.account)
        if form.is_valid():
            line_item = form.save(commit=False)
            line_item.estimate = estimate
            line_item.source = LineItem.Source.MANUAL
            line_item.save()
            messages.success(request, f'Added {line_item.quantity} x {line_item.material.name}.')
        else:
            errors = '; '.join(
                f'{field}: {", ".join(field_errors)}' for field, field_errors in form.errors.items()
            )
            messages.error(request, f'Could not add line: {errors}')
        return redirect('estimating:estimate-detail', pk=estimate.pk)


class ManualLineItemDeleteView(LoginRequiredMixin, View):
    """Manual lines only. Tool-generated lines are owned by their trace - they
    would silently reappear on the next regeneration, so deleting them here
    would be misleading. Remove those by deleting or editing the trace."""

    def post(self, request, pk):
        line_item = get_object_or_404(
            LineItem.objects.filter(
                estimate__project__account=request.user.account,
                source=LineItem.Source.MANUAL,
            ),
            pk=pk,
        )
        estimate_id = line_item.estimate_id
        line_item.delete()
        messages.success(request, 'Line removed.')
        return redirect('estimating:estimate-detail', pk=estimate_id)


class MaterialImportView(LoginRequiredMixin, View):
    """Import supplier material lists (CSV or XLSX) into the account's own
    catalog, e.g. a manufacturer's hanger list. Header row required; columns
    (case-insensitive): name (required), category (key or label), species,
    grade, dimension, input_type (ft/box/each), quantity_per_box,
    lengths (semicolon/comma separated feet), default_length. Rows whose name
    already exists for this account are skipped, so re-importing an updated
    file is safe."""

    template_name = 'estimating/material_import.html'
    MAX_BYTES = 2 * 1024 * 1024
    MAX_ROWS = 2000

    def get(self, request):
        return render(request, self.template_name)

    def post(self, request):
        from decimal import InvalidOperation

        from catalog.models import MaterialLength, MaterialProduct

        upload = request.FILES.get('file')
        if upload is None:
            messages.error(request, 'Choose a CSV or XLSX file to import.')
            return render(request, self.template_name)
        if upload.size > self.MAX_BYTES:
            messages.error(request, 'File is too large (2 MB limit).')
            return render(request, self.template_name)

        try:
            rows = self._read_rows(upload)
        except ValueError as exc:
            messages.error(request, str(exc))
            return render(request, self.template_name)

        label_to_key = {label.lower(): key for key, label in MaterialProduct.Category.choices}
        valid_keys = {key for key, _ in MaterialProduct.Category.choices}
        account = request.user.account
        existing = set(
            MaterialProduct.objects.filter(account=account).values_list('name', flat=True)
        )

        created, skipped, errors = 0, 0, []
        with transaction.atomic():
            for line_number, row in rows:
                name = (row.get('name') or '').strip()
                if not name:
                    errors.append(f'Row {line_number}: missing name.')
                    continue
                if name in existing:
                    skipped += 1
                    continue

                raw_category = (row.get('category') or '').strip().lower()
                category = (
                    raw_category if raw_category in valid_keys
                    else label_to_key.get(raw_category, MaterialProduct.Category.UNCATEGORIZED)
                )
                input_type = (row.get('input_type') or 'each').strip().lower()
                if input_type not in ('ft', 'box', 'each'):
                    errors.append(f'Row {line_number}: input_type must be ft, box, or each.')
                    continue

                quantity_per_box = None
                if input_type == 'box':
                    try:
                        quantity_per_box = int(float(row.get('quantity_per_box') or 0)) or None
                    except (TypeError, ValueError):
                        quantity_per_box = None
                    if quantity_per_box is None:
                        errors.append(f'Row {line_number}: box materials need quantity_per_box.')
                        continue

                lengths, default_length = [], None
                if input_type == 'ft':
                    raw_lengths = str(row.get('lengths') or '').replace(';', ',')
                    try:
                        lengths = [Decimal(part.strip()) for part in raw_lengths.split(',') if part.strip()]
                        raw_default = str(row.get('default_length') or '').strip()
                        default_length = Decimal(raw_default) if raw_default else None
                    except (InvalidOperation, ValueError):
                        errors.append(f'Row {line_number}: lengths must be numbers of feet.')
                        continue
                    if not lengths:
                        errors.append(f'Row {line_number}: ft materials need at least one stock length.')
                        continue
                    if default_length is None or default_length not in lengths:
                        default_length = lengths[0]

                product = MaterialProduct.objects.create(
                    account=account, name=name, slug=slugify(name),
                    category=category,
                    species=(row.get('species') or '').strip(),
                    grade=(row.get('grade') or '').strip(),
                    nominal_dimension=(row.get('dimension') or '').strip(),
                    supported_input_types=[input_type],
                    input_type=input_type, quantity_per_box=quantity_per_box,
                )
                for length in lengths:
                    MaterialLength.objects.create(
                        product=product, length_ft=length, is_default=(length == default_length),
                    )
                existing.add(name)
                created += 1

        summary = f'Imported {created} material{"s" if created != 1 else ""}.'
        if skipped:
            summary += f' Skipped {skipped} already in your library.'
        messages.success(request, summary)
        for error in errors[:10]:
            messages.warning(request, error)
        if len(errors) > 10:
            messages.warning(request, f'...and {len(errors) - 10} more rows had problems.')
        return redirect('estimating:library')

    def _read_rows(self, upload):
        """Yields (line_number, {header: value}) for CSV or XLSX uploads."""
        name = (upload.name or '').lower()
        if name.endswith('.xlsx'):
            try:
                from openpyxl import load_workbook
            except ImportError as exc:
                raise ValueError('XLSX support requires the openpyxl package.') from exc
            try:
                sheet = load_workbook(upload, read_only=True, data_only=True).worksheets[0]
            except Exception as exc:
                raise ValueError('Could not read that XLSX file.') from exc
            iterator = sheet.iter_rows(values_only=True)
            try:
                headers = [str(cell or '').strip().lower() for cell in next(iterator)]
            except StopIteration:
                raise ValueError('The file is empty.') from None
            rows = []
            for index, values in enumerate(iterator, start=2):
                if index - 1 > self.MAX_ROWS:
                    raise ValueError(f'Too many rows (limit {self.MAX_ROWS}).')
                row = {headers[i]: values[i] for i in range(min(len(headers), len(values)))}
                if any(str(v or '').strip() for v in row.values()):
                    rows.append((index, row))
            return rows
        if name.endswith('.csv'):
            import io
            try:
                text = upload.read().decode('utf-8-sig')
            except UnicodeDecodeError:
                raise ValueError('CSV must be UTF-8 encoded.') from None
            reader = csv.DictReader(io.StringIO(text))
            if not reader.fieldnames:
                raise ValueError('The file is empty.')
            reader.fieldnames = [(field or '').strip().lower() for field in reader.fieldnames]
            rows = []
            for index, row in enumerate(reader, start=2):
                if index - 1 > self.MAX_ROWS:
                    raise ValueError(f'Too many rows (limit {self.MAX_ROWS}).')
                if any((value or '').strip() for value in row.values() if isinstance(value, str)):
                    rows.append((index, row))
            return rows
        raise ValueError('Upload a .csv or .xlsx file.')


class MaterialCreateView(LoginRequiredMixin, View):
    template_name = 'estimating/material_form.html'

    def get(self, request):
        material = self._material_for_request(request)
        return self._render(request, MaterialForm(instance=material, account=request.user.account), created=True)

    def post(self, request):
        material = self._material_for_request(request)
        form = MaterialForm(request.POST, instance=material, account=request.user.account)
        if form.is_valid():
            material = form.save()
            messages.success(request, f'Material "{material.name}" created.')
            return redirect('estimating:library')
        return self._render(request, form, created=True)

    def _material_for_request(self, request):
        from catalog.models import MaterialProduct

        return MaterialProduct(account=request.user.account)

    def _render(self, request, form, *, created):
        return render(request, self.template_name, {
            'form': form,
            'material': form.instance,
            'created': created,
        })


class MaterialUpdateView(LoginRequiredMixin, View):
    template_name = 'estimating/material_form.html'

    def get(self, request, pk):
        material = self._get_material(request, pk)
        return self._render(request, MaterialForm(instance=material, account=request.user.account), created=False)

    def post(self, request, pk):
        material = self._get_material(request, pk)
        form = MaterialForm(request.POST, instance=material, account=request.user.account)
        if form.is_valid():
            material = form.save()
            messages.success(request, f'Material "{material.name}" updated.')
            return redirect('estimating:library')
        return self._render(request, form, created=False)

    def _get_material(self, request, pk):
        from catalog.models import MaterialProduct

        return get_object_or_404(
            MaterialProduct.objects.filter(account=request.user.account).prefetch_related('lengths'),
            pk=pk,
        )

    def _render(self, request, form, *, created):
        return render(request, self.template_name, {
            'form': form,
            'material': form.instance,
            'created': created,
        })


class MaterialPriceUpdateView(LoginRequiredMixin, View):
    """Sets the account's private unit cost for any material the account can
    see, including global stock SKUs. Global material attributes stay
    read-only (edited only through MaterialUpdateView on owned copies), but
    pricing is always account-scoped, so pricing a stock 2x6 or a seeded
    hanger is safe and never mutates the shared catalog. A blank cost clears
    the price."""

    def post(self, request, pk):
        from decimal import InvalidOperation

        from catalog.models import MaterialPrice, MaterialProduct

        material = get_object_or_404(MaterialProduct.objects.visible_to(request.user.account), pk=pk)
        raw = (request.POST.get('unit_cost') or '').strip()
        if not raw:
            MaterialPrice.objects.filter(account=request.user.account, material=material).delete()
            messages.success(request, f'Cleared the price for {material.name}.')
            return redirect('estimating:library')
        try:
            unit_cost = Decimal(raw)
        except (InvalidOperation, ValueError):
            messages.error(request, 'Enter a valid price, e.g. 7.50.')
            return redirect('estimating:library')
        if unit_cost < 0:
            messages.error(request, 'Price cannot be negative.')
            return redirect('estimating:library')
        MaterialPrice.objects.update_or_create(
            account=request.user.account, material=material,
            defaults={'unit_cost': unit_cost},
        )
        messages.success(request, f'Set {material.name} to ${unit_cost:.2f}.')
        return redirect('estimating:library')


class MaterialDeleteView(LoginRequiredMixin, View):
    def post(self, request, pk):
        from catalog.models import MaterialProduct

        material = get_object_or_404(MaterialProduct.objects.filter(account=request.user.account), pk=pk)
        name = material.name
        try:
            material.delete()
        except ProtectedError:
            messages.error(
                request,
                f'Could not delete "{name}" because it is still used by an assembly or estimate line.',
            )
        else:
            messages.success(request, f'Material "{name}" deleted.')
        return redirect('estimating:library')


class AssemblyQuickEditView(LoginRequiredMixin, View):
    """Powers the Library's quick-edit drawer: GET returns an assembly's rules
    as JSON; POST applies per-rule material and waste changes.

    Copy-on-write: a global (seeded) assembly is never mutated. Posting edits
    against one clones it into the account's library as '<name> (Custom)' with
    the changes applied, so every other tenant keeps the stock version.
    Posting against an already-owned assembly updates it in place. Repeat
    edits against the same global assembly reuse the existing clone."""

    def _get_assembly(self, request, pk):
        return get_object_or_404(
            Assembly.objects.visible_to(request.user.account).prefetch_related('rules__material'),
            pk=pk,
        )

    def get(self, request, pk):
        assembly = self._get_assembly(request, pk)
        return JsonResponse({
            'id': assembly.pk,
            'name': assembly.name,
            'is_global': assembly.account_id is None,
            'tool_type': assembly.tool_type,
            'rules': [
                {
                    'id': rule.pk,
                    'role': rule.role,
                    'kind': (rule.formula.name if rule.formula_id else rule.get_formula_kind_display()),
                    'material_id': rule.material_id,
                    'waste_factor': str(rule.waste_factor),
                }
                for rule in assembly.rules.all().order_by('order')
            ],
        })

    def post(self, request, pk):
        from catalog.models import MaterialProduct

        account = request.user.account
        assembly = self._get_assembly(request, pk)

        try:
            payload = json.loads(request.body)
            edits = {int(rule['id']): rule for rule in payload['rules']}
        except (json.JSONDecodeError, KeyError, TypeError, ValueError):
            return JsonResponse({'error': 'Invalid payload.'}, status=400)

        # Validate every edit up front against the source assembly's rules.
        source_rules = {rule.pk: rule for rule in assembly.rules.all()}
        visible_material_ids = set(
            MaterialProduct.objects.visible_to(account).values_list('id', flat=True)
        )
        validated = {}
        for rule_id, edit in edits.items():
            if rule_id not in source_rules:
                return JsonResponse({'error': 'Unknown rule for this assembly.'}, status=400)
            try:
                material_id = int(edit['material_id'])
                waste = Decimal(str(edit['waste_factor']))
            except (KeyError, TypeError, ValueError, ArithmeticError):
                return JsonResponse({'error': 'Invalid rule edit.'}, status=400)
            if material_id not in visible_material_ids:
                return JsonResponse({'error': 'Invalid material.'}, status=400)
            if not (Decimal('0') <= waste <= Decimal('1')):
                return JsonResponse({'error': 'Waste factor must be between 0 and 100 percent.'}, status=400)
            validated[rule_id] = {'material_id': material_id, 'waste_factor': waste}

        with transaction.atomic():
            if assembly.account_id is None:
                saved, cloned = self._apply_to_clone(account, assembly, validated)
            else:
                saved, cloned = assembly, False
                for rule_id, edit in validated.items():
                    rule = source_rules[rule_id]
                    rule.material_id = edit['material_id']
                    rule.waste_factor = edit['waste_factor']
                    rule.save(update_fields=['material', 'waste_factor'])

        return JsonResponse({'id': saved.pk, 'name': saved.name, 'cloned': cloned})

    def _apply_to_clone(self, account, source, validated):
        """Find or create this account's clone of a global assembly, then apply
        the validated edits. Clone rules are matched back to source rules by
        (order, role), which the clone copies verbatim."""
        clone_name = f'{source.name} (Custom)'
        clone = Assembly.objects.filter(
            account=account, tool_type=source.tool_type, name=clone_name,
        ).first()
        # A second lookup guards the (account, opening_kind, wall_subtype)
        # uniqueness: an opening assembly cloned under a different name earlier
        # must be reused, not duplicated.
        if clone is None and source.opening_kind and source.wall_subtype:
            clone = Assembly.objects.filter(
                account=account, opening_kind=source.opening_kind, wall_subtype=source.wall_subtype,
            ).first()

        created = clone is None
        if created:
            clone = Assembly.objects.create(
                account=account, name=clone_name, tool_type=source.tool_type,
                category=source.category, wall_subtype=source.wall_subtype,
                opening_kind=source.opening_kind, beam_type=source.beam_type,
                description=source.description,
                # Never default: it would sit beside the global default in the
                # viewer's candidate set and make auto-select ambiguous. Tool
                # memory re-selects it after the first manual pick anyway.
                is_default=False,
            )
            for rule in source.rules.all().order_by('order'):
                edit = validated.get(rule.pk, {})
                CalculationRule.objects.create(
                    assembly=clone, formula=rule.formula, role=rule.role,
                    formula_kind=rule.formula_kind, multiplier=rule.multiplier,
                    extra=rule.extra, coverage_sqft=rule.coverage_sqft,
                    units_per_measurement=rule.units_per_measurement,
                    corner_stud_count=rule.corner_stud_count,
                    t_intersection_stud_count=rule.t_intersection_stud_count,
                    t_backer_stud_count=rule.t_backer_stud_count,
                    order=rule.order,
                    material_id=edit.get('material_id', rule.material_id),
                    waste_factor=edit.get('waste_factor', rule.waste_factor),
                )
        else:
            clone_rules = {(r.order, r.role): r for r in clone.rules.all()}
            for rule_id, edit in validated.items():
                source_rule = CalculationRule.objects.get(pk=rule_id)
                target = clone_rules.get((source_rule.order, source_rule.role))
                if target is not None:
                    target.material_id = edit['material_id']
                    target.waste_factor = edit['waste_factor']
                    target.save(update_fields=['material', 'waste_factor'])
        return clone, created


class LibraryView(LoginRequiredMixin, TemplateView):
    template_name = 'estimating/library.html'

    def get_context_data(self, **kwargs):
        from itertools import groupby

        from catalog.models import MaterialProduct, MaterialPrice

        context = super().get_context_data(**kwargs)
        account = self.request.user.account

        # Materials, alphabetized within each category, categories in label order.
        materials = list(
            MaterialProduct.objects.visible_to(account)
            .prefetch_related('lengths')
            .order_by('category', 'name')
        )
        price_map = dict(
            MaterialPrice.objects.filter(account=account)
            .values_list('material_id', 'unit_cost')
        )
        for material in materials:
            material.library_unit_cost = price_map.get(material.id)
            material.library_length_values = list(material.lengths.all())
            material.library_default_length_ft = None
            material.library_supported_inputs = material.supported_input_type_labels
            material.library_supports_ft = material.supports_input_type(MaterialProduct.InputType.FT)
            material.library_supports_box = material.supports_input_type(MaterialProduct.InputType.BOX)
            material.library_supports_each = material.supports_input_type(MaterialProduct.InputType.EACH)
            if material.library_supports_ft:
                default = next((length for length in material.library_length_values if length.is_default), None)
                if default is not None:
                    material.library_default_length_ft = default.length_ft
        label_for = dict(MaterialProduct.Category.choices)
        material_groups = []
        for category_key, items in groupby(materials, key=lambda m: m.category):
            material_groups.append({
                'key': category_key,
                'label': label_for.get(category_key, 'Uncategorized'),
                'materials': list(items),
            })
        material_groups.sort(key=lambda g: g['label'])

        # Assemblies grouped by construction-system category.
        assemblies = list(
            Assembly.objects.visible_to(account)
            .prefetch_related('rules__material', 'rules__formula')
            .order_by('category', 'name')
        )
        assembly_label = dict(Assembly.Category.choices)
        assembly_groups = []
        for category_key, items in groupby(assemblies, key=lambda a: a.category):
            assembly_groups.append({
                'key': category_key or 'uncategorized',
                'label': assembly_label.get(category_key, 'Uncategorized'),
                'assemblies': sorted(items, key=lambda a: a.name.lower()),
            })

        context['material_groups'] = material_groups
        context['material_count'] = len(materials)
        # Flat alphabetized list for the quick-edit drawer's material selects.
        context['materials_flat'] = sorted(
            ({'id': m.id, 'name': m.name} for m in materials), key=lambda m: m['name'].lower(),
        )
        context['assembly_groups'] = assembly_groups
        context['assembly_count'] = len(assemblies)
        context['formulas'] = (
            Formula.objects.visible_to(account)
            .select_related('base_formula').order_by('name')
        )
        return context


class FormulaCreateView(LoginRequiredMixin, CreateView):
    form_class = FormulaForm
    template_name = 'estimating/formula_form.html'
    success_url = reverse_lazy('estimating:library')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['account'] = self.request.user.account
        return kwargs

    def form_valid(self, form):
        messages.success(self.request, f'Formula “{form.cleaned_data["name"]}” created.')
        return super().form_valid(form)


class AssemblyCreateView(LoginRequiredMixin, View):
    template_name = 'estimating/assembly_form.html'

    def get(self, request):
        assembly = Assembly(account=request.user.account)
        return self._render(request, AssemblyForm(instance=assembly), CalculationRuleFormSet(
            instance=assembly, form_kwargs={'account': request.user.account},
        ))

    def post(self, request):
        assembly = Assembly(account=request.user.account)
        form = AssemblyForm(request.POST, instance=assembly)
        rules = CalculationRuleFormSet(
            request.POST, instance=assembly, form_kwargs={'account': request.user.account},
        )
        if form.is_valid() and rules.is_valid():
            with transaction.atomic():
                assembly = form.save()
                rules.instance = assembly
                rules.save()
            messages.success(request, f'Assembly “{assembly.name}” created.')
            return redirect('estimating:library')
        return self._render(request, form, rules)

    def _render(self, request, form, rules):
        return render(request, self.template_name, {'form': form, 'rules': rules})
